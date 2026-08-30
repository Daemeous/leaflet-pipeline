"""
add_unnamed_roads.py
=====================
Adds the unnamed-road clusters from unnamed_roads_report.csv into
<prefix>_Leafletting_residences_uprn.xlsx as placeholder "Unknown Road" rows,
appended at the end, so they render on the map and can be visually spotted /
named by hand later (Street is just a text cell — real names can replace
"Unknown Road" without touching geometry).

Reproduces the exact same clustering as fill_gaps.py (same cached
all_highways_raw.json => deterministic) to recover each cluster's full
geometry (the CSV only kept a centroid), assigns each a dominant ward the
same way run_pipeline.py does (length-in-ward), and estimates Residences
via a standalone 40m-buffer UPRN count (NOT jointly disambiguated against
the other named roads — a UPRN near both a named road and one of these
could in principle be counted by both; fine for this exploratory pass).

Input/Output: <prefix>_Leafletting_residences_uprn.xlsx, updated in place
with the unnamed clusters appended as new rows.
"""

import json
import time
from collections import Counter, defaultdict

import openpyxl
import pandas as pd
import geopandas as gpd
from shapely.geometry import LineString
from shapely.ops import unary_union

from constituency_config import get_config
from fill_gaps import ALL_HIGHWAYS_JSON, WARDS_GEOJSON, LEAFLETABLE_TYPES, ROAD_LIKE_TYPES, _DISTRICT_BY_WARD
from run_pipeline import _geometry_to_text

_cfg = get_config()
_prefix = _cfg["output_prefix"]

TARGET_XLSX = f"{_prefix}_Leafletting_residences_uprn.xlsx"
CONSTITUENCY_GEOJSON = f"{_prefix}_constituency.geojson"
UPRN_CSV = "../data/osopenuprn_202605.csv"
ROAD_BUFFER_METRES = 40


def main():
    print(f"Loading {ALL_HIGHWAYS_JSON} ...")
    with open(ALL_HIGHWAYS_JSON, encoding="utf-8") as f:
        osm = json.load(f)
    elements = [e for e in osm["elements"] if e.get("type") == "way"]

    named_elements = [e for e in elements if e.get("tags", {}).get("name", "").strip()]
    unnamed_elements = [e for e in elements if not e.get("tags", {}).get("name", "").strip()]

    node_to_names = defaultdict(set)
    for el in named_elements:
        name = el["tags"]["name"].strip()
        for nid in el.get("nodes", []):
            node_to_names[nid].add(name)

    standalone = []
    for el in unnamed_elements:
        nodes = el.get("nodes", [])
        pts = el.get("geometry", [])
        if len(nodes) < 2 or len(pts) < 2:
            continue
        ht = el.get("tags", {}).get("highway")
        endpoint_names = set()
        if ht in ROAD_LIKE_TYPES:
            for nid in (nodes[0], nodes[-1]):
                endpoint_names |= node_to_names.get(nid, set())
        if len(endpoint_names) == 1:
            continue  # gap-fill absorbed this one already
        elif ht in LEAFLETABLE_TYPES:
            standalone.append(el)

    wards_gdf = gpd.read_file(WARDS_GEOJSON).to_crs(4326)
    constituency_union = unary_union(list(wards_gdf.geometry))
    standalone = [
        el for el in standalone
        if LineString([(p["lon"], p["lat"]) for p in el.get("geometry", [])]).intersects(constituency_union)
    ]

    # ── Same clustering as fill_gaps.py ───────────────────────────────────────
    parent = {el["id"]: el["id"] for el in standalone}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    node_to_ids = defaultdict(list)
    for el in standalone:
        for nid in el.get("nodes", []):
            node_to_ids[nid].append(el["id"])
    for ids in node_to_ids.values():
        for other in ids[1:]:
            union(ids[0], other)

    clusters = defaultdict(list)
    for el in standalone:
        clusters[find(el["id"])].append(el)

    MAX_CLUSTER_LENGTH_M = 800
    final_clusters = []
    for els in clusters.values():
        length = sum(
            LineString([(p["lon"], p["lat"]) for p in el.get("geometry", [])]).length * 111_320
            for el in els
        )
        if length <= MAX_CLUSTER_LENGTH_M or len(els) == 1:
            final_clusters.append(els)
        else:
            final_clusters.extend([el] for el in els)

    print(f"{len(final_clusters):,} unnamed clusters "
          f"(should match unnamed_roads_report.csv row count)")

    # ── Ward assignment: dominant ward by length-in-ward, same as run_pipeline.py
    ward_records = []
    for _, row in wards_gdf.iterrows():
        gpkg_name = row["Name"]
        display = gpkg_name[:-5] if gpkg_name.endswith(" Ward") else gpkg_name
        district = _DISTRICT_BY_WARD.get(display.strip().lower(), "Unknown")
        ward_records.append((display, district, row.geometry))

    cluster_rows = []
    for els in final_clusters:
        segments = [LineString([(p["lon"], p["lat"]) for p in el.get("geometry", [])]) for el in els]
        merged = unary_union(segments)
        total_len = merged.length
        if total_len == 0:
            continue
        best_ward, best_district, best_len = None, "Unknown", 0
        for display, district, poly in ward_records:
            clipped = merged.intersection(poly)
            if clipped.is_empty:
                continue
            if clipped.length > best_len:
                best_len, best_ward, best_district = clipped.length, display, district
        if best_ward is None:
            centroid = merged.centroid
            best_ward, best_district, _ = min(ward_records, key=lambda wr: wr[2].distance(centroid))
        lat, lon = round(merged.centroid.y, 6), round(merged.centroid.x, 6)
        cluster_rows.append({
            "Street": "Unknown Road", "@lat": lat, "@lon": lon,
            "Ward": best_ward, "Local Authority District": best_district,
            "Status": "Not_Started", "Residences": 0,
            "road_geometry": _geometry_to_text(merged), "partial_geometry": "-",
            "_geom": merged,
        })

    print("Ward breakdown of the unnamed clusters:")
    print(Counter(r["Ward"] for r in cluster_rows))

    # ── Standalone 40m-buffer UPRN count (not disambiguated against named roads)
    print(f"\nLoading UPRNs from {UPRN_CSV} (filtered to constituency bbox) ...")
    boundary = gpd.read_file(CONSTITUENCY_GEOJSON)
    constituency_boundary = boundary.geometry.iloc[0]
    constituency_padded = constituency_boundary.buffer(0.002)
    bbox = constituency_padded.bounds

    t0 = time.time()
    kept = []
    for chunk in pd.read_csv(
        UPRN_CSV, chunksize=500_000,
        usecols=["UPRN", "LATITUDE", "LONGITUDE"],
        dtype={"UPRN": "int64", "LATITUDE": "float64", "LONGITUDE": "float64"},
    ):
        pre = chunk[
            chunk["LONGITUDE"].between(bbox[0], bbox[2]) &
            chunk["LATITUDE"].between(bbox[1], bbox[3])
        ]
        if len(pre) > 0:
            pts = gpd.GeoDataFrame(pre, geometry=gpd.points_from_xy(pre["LONGITUDE"], pre["LATITUDE"]), crs="EPSG:4326")
            inside = pts[pts.geometry.within(constituency_padded)]
            kept.append(inside[["UPRN", "geometry"]])
    uprn_gdf = pd.concat(kept, ignore_index=True)
    print(f"  {len(uprn_gdf):,} UPRNs in constituency [{time.time()-t0:.0f}s]")

    uprn_proj = gpd.GeoDataFrame(uprn_gdf, geometry="geometry", crs="EPSG:4326").to_crs("EPSG:27700")

    clusters_gdf = gpd.GeoDataFrame(cluster_rows, geometry=[r["_geom"] for r in cluster_rows], crs="EPSG:4326")
    clusters_proj = clusters_gdf.to_crs("EPSG:27700")
    clusters_buffered = clusters_proj.copy()
    clusters_buffered["geometry"] = clusters_proj.geometry.buffer(ROAD_BUFFER_METRES)

    joined = gpd.sjoin(
        uprn_proj[["UPRN", "geometry"]],
        clusters_buffered[["geometry"]].reset_index().rename(columns={"index": "cluster_idx"}),
        how="inner", predicate="within",
    )
    counts = defaultdict(int)
    for uprn_i, group in joined.groupby(level=0):
        if len(group) == 1:
            counts[int(group["cluster_idx"].iloc[0])] += 1
        else:
            pt = uprn_proj.loc[uprn_i, "geometry"]
            best = min(group["cluster_idx"], key=lambda ci: pt.distance(clusters_proj.loc[ci, "geometry"]))
            counts[int(best)] += 1

    for idx, row in enumerate(cluster_rows):
        row["Residences"] = counts.get(idx, 0)
        del row["_geom"]

    print(f"\nTotal residences assigned to unnamed clusters: {sum(r['Residences'] for r in cluster_rows):,}")
    print(f"Clusters with >=1 residence: {sum(1 for r in cluster_rows if r['Residences'] > 0)} of {len(cluster_rows)}")

    # ── Append to the corrected road set ──────────────────────────────────────
    print(f"\nAppending to {TARGET_XLSX} ...")
    wb = openpyxl.load_workbook(TARGET_XLSX, data_only=False)
    ws = wb["Data"]
    headers = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}

    append_row = ws.max_row + 1
    for r in sorted(cluster_rows, key=lambda x: x["Ward"]):
        ws.cell(append_row, col["Street"]).value = r["Street"]
        ws.cell(append_row, col["@lat"]).value = r["@lat"]
        ws.cell(append_row, col["@lon"]).value = r["@lon"]
        ws.cell(append_row, col["Ward"]).value = r["Ward"]
        ws.cell(append_row, col["Local Authority District"]).value = r["Local Authority District"]
        ws.cell(append_row, col["Status"]).value = r["Status"]
        ws.cell(append_row, col["Residences"]).value = r["Residences"]
        ws.cell(append_row, col["road_geometry"]).value = r["road_geometry"]
        if "partial_geometry" in col:
            ws.cell(append_row, col["partial_geometry"]).value = "-"
        append_row += 1

    wb.save(TARGET_XLSX)
    print(f"Appended {len(cluster_rows)} 'Unknown Road' rows. New Data row count: {ws.max_row - 1}")
    print(f"Saved: {TARGET_XLSX}")


if __name__ == "__main__":
    main()
