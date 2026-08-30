"""
finalize_output.py
====================
Folds the corrected/gap-filled/UPRN'd road set (<prefix>_Leafletting_residences_uprn.xlsx,
including the appended "Unknown Road" placeholder rows) into a copy of
BASE_XLSX — the existing reference sheet for this constituency — producing
OUTPUT_XLSX ready to review and paste back into whatever's actually live.

Only use this for REFRESHING an already-built constituency, never for the
initial build (there's nothing to merge into yet the first time round).

Rules as written below assume the base sheet has no real tracked progress
worth protecting (check first — `df['Status'].value_counts()` on BASE_XLSX;
if it's not 100% Not_Started, a coordinator has been using this sheet and
a blind "every column but Status" overwrite could silently change a
residence count attached to a road someone already marked Complete/
In_Progress. If that's the case, add a status-aware gate before applying
row-in-place updates — don't just run this script as-is; see the
[[stafford-missing-roads-fix]] memory for the more cautious version that
gates on Status and writes a review CSV instead of touching those rows):
  - Existing rows are matched to fresh data by (Street, Ward), case-insensitive.
    Every column EXCEPT Status is overwritten with the fresh value.
  - Rows in the base with no fresh match at all are left completely alone
    and logged to review_no_fresh_match.csv (likely renamed/reassigned by
    the clustering fix — worth a manual look, never auto-deleted).
  - Fresh (Street, Ward) pairs not present in the base are APPENDED at the
    end, in the same row order every time it's regenerated — never
    inserted/reordered — so row positions stay stable if this is re-run.
  - Every other sheet in BASE_XLSX (if any) is copied through untouched.

Set BASE_XLSX / OUTPUT_XLSX below before running.
"""

import openpyxl
import pandas as pd

from constituency_config import get_config

_cfg = get_config()
_prefix = _cfg["output_prefix"]

FRESH_XLSX = f"{_prefix}_Leafletting_residences_uprn.xlsx"

# ── Fill these in per constituency before running ─────────────────────────
BASE_XLSX = "REPLACE_ME.xlsx"
OUTPUT_XLSX = f"{_prefix}_UPDATED.xlsx"
# ────────────────────────────────────────────────────────────────────────


def norm(s):
    return str(s).strip().lower()


def main():
    print(f"Loading fresh data from {FRESH_XLSX} ...")
    wb_fresh = openpyxl.load_workbook(FRESH_XLSX, data_only=True)
    ws_fresh = wb_fresh["Data"]
    fresh = {}
    for r in ws_fresh.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        street, lat, lon, ward, lad, status, resid, geom, partial = r
        fresh[(norm(street), norm(ward))] = dict(
            street=street.strip(), lat=lat, lon=lon, ward=ward.strip(),
            lad=lad, resid=resid, geom=geom,
        )
    print(f"  {len(fresh):,} fresh rows")

    print(f"Loading base sheet from {BASE_XLSX} (full workbook, formulas preserved) ...")
    wb_base = openpyxl.load_workbook(BASE_XLSX, data_only=False)
    ws_base = wb_base["Data"]

    headers = [c.value for c in ws_base[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}
    print("Data columns:", headers)

    last_row = ws_base.max_row
    current = {}
    for row_num in range(2, last_row + 1):
        street = ws_base.cell(row_num, col["Street"]).value
        ward = ws_base.cell(row_num, col["Ward"]).value
        if not street:
            continue
        current[(norm(street), norm(ward))] = row_num

    n_updated = n_appended = 0
    review_no_match = []
    matched_fresh_keys = set()

    for key, row_num in current.items():
        if key not in fresh:
            review_no_match.append((
                ws_base.cell(row_num, col["Street"]).value,
                ws_base.cell(row_num, col["Ward"]).value,
                ws_base.cell(row_num, col["Status"]).value,
                ws_base.cell(row_num, col["Residences"]).value,
            ))
            continue

        matched_fresh_keys.add(key)
        f = fresh[key]
        ws_base.cell(row_num, col["Street"]).value = f["street"]
        ws_base.cell(row_num, col["@lat"]).value = f["lat"]
        ws_base.cell(row_num, col["@lon"]).value = f["lon"]
        ws_base.cell(row_num, col["Ward"]).value = f["ward"]
        ws_base.cell(row_num, col["Local Authority District"]).value = f["lad"]
        ws_base.cell(row_num, col["Residences"]).value = f["resid"]
        ws_base.cell(row_num, col["road_geometry"]).value = f["geom"]
        if "partial_geometry" in col:
            ws_base.cell(row_num, col["partial_geometry"]).value = "-"
        n_updated += 1

    new_rows = [f for key, f in fresh.items() if key not in matched_fresh_keys]

    append_row = last_row + 1
    for f in sorted(new_rows, key=lambda x: (x["ward"], x["street"])):
        ws_base.cell(append_row, col["Street"]).value = f["street"]
        ws_base.cell(append_row, col["@lat"]).value = f["lat"]
        ws_base.cell(append_row, col["@lon"]).value = f["lon"]
        ws_base.cell(append_row, col["Ward"]).value = f["ward"]
        ws_base.cell(append_row, col["Local Authority District"]).value = f["lad"]
        ws_base.cell(append_row, col["Status"]).value = "Not_Started"
        ws_base.cell(append_row, col["Residences"]).value = f["resid"]
        ws_base.cell(append_row, col["road_geometry"]).value = f["geom"]
        if "partial_geometry" in col:
            ws_base.cell(append_row, col["partial_geometry"]).value = "-"
        append_row += 1
        n_appended += 1

    wb_base.save(OUTPUT_XLSX)

    print()
    print("=" * 60)
    print("MERGE COMPLETE")
    print("=" * 60)
    print(f"Updated in place (all columns but Status): {n_updated}")
    print(f"Appended as new rows:                       {n_appended}")
    print(f"Flagged — base row with no fresh match:      {len(review_no_match)}")
    print(f"\nSaved: {OUTPUT_XLSX}")

    if review_no_match:
        df = pd.DataFrame(review_no_match, columns=["Street", "Ward", "Status", "Residences"])
        df.to_csv("review_no_fresh_match.csv", index=False)
        print(f"Wrote review_no_fresh_match.csv ({len(df)} rows)")


if __name__ == "__main__":
    main()
