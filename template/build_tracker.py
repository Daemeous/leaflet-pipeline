"""
build_tracker.py  (TEMPLATE)
=============================
Turns <prefix>_Leafletting_residences_uprn.xlsx (from estimate_residences_uprn.py)
into a Google-Sheets-ready tracker workbook: <prefix>_Tracker.xlsx.

This mirrors the sheet structure of an existing hand-built tracker
(Barnsley, Penistone & Stocksbridge.xlsx) that's already hosted as a Google
Sheet with an Apps Script bound to it for edit-logging and access control.
That reference file had a real bug worth knowing about: its Dashboard sheet
was copy-pasted from an earlier, smaller (18-ward) constituency and never
had its "Overall" row / summary formulas re-ranged for the new ward count —
they silently summed the wrong rows. This script avoids that whole bug
class by generating every range from the actual ward list and actual Data
row count, every time, rather than hand-editing a copy.

Adds four sheets around the existing Data sheet:
  - Changelog   header row only; an Apps Script bound to the Google Sheet
                logs status edits here (timestamp, rowIndex, street, ward,
                field, oldValue, newValue, editorEmail, isRevert). Column
                names/order must NOT change — the Apps Script maps by name.
  - Dashboard   one row per ward (COUNTIF/SUMIF formulas driven off Data),
                an Overall total row, a Borough Summary KPI block, a
                Residences Remaining column, and four charts (two ported
                from the reference file, two new — see below).
  - Checksum    one formula the Apps Script can poll to detect edits made
                outside its own logging path.
  - Authorised  email addresses allowed to bypass the Apps Script's edit
                locks — from constituency_config.py's `authorised_emails`.

Also finishes formatting the Data sheet: freezes the header row, hides the
@lat column (kept for an Apps Script map-link formula, not for manual
reading — matches the reference file), and adds the Status dropdown
(Not_Started / Planned / In_Progress / Complete) the Apps Script and
colour-coding both rely on.

New in this Dashboard vs. the reference file (useful additions, not just a
faithful copy):
  - "Residences Remaining" column (H - I) with its own colour scale, so the
    wards with the most work left jump out without doing mental subtraction.
  - A proper "Borough Summary" KPI block replacing two barely-labelled
    formulas — total roads/residences, residences reached (complete-only
    and weighted-with-in-progress), residences remaining, % reached, and a
    ward-completion headcount.
  - A borough-wide status donut, so the overall picture doesn't have to be
    picked out as one bar among many in the per-ward chart.
  - A "Residences Remaining by Ward" bar chart, a direct visual prioritiser
    for where the next leaflet run has the most impact.

Run this LAST, after estimate_residences_uprn.py.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import AreaChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from constituency_config import get_config

cfg = get_config()
_prefix = cfg["output_prefix"]

EXCEL_INPUT  = f"{_prefix}_Leafletting_residences_uprn.xlsx"
EXCEL_OUTPUT = f"{_prefix}_Tracker.xlsx"

STATUS_LIST = "Not_Started,Planned,In_Progress,Complete"

DASHBOARD_HEADERS = [
    "Ward", "Streets", "Complete", "In Progress", "Planned", "Not Started",
    "% Roads Complete", "Residences", "Estimate served", "% Residences Complete",
    "Overseer", "Residences Remaining",
]

CHANGELOG_HEADERS = [
    "timestamp", "rowIndex", "street", "ward", "field",
    "oldValue", "newValue", "editorEmail", "isRevert",
]


def format_data_sheet(ws, n_data_rows):
    ws.freeze_panes = "A2"
    ws.column_dimensions["B"].hidden = True

    widths = {"A": 37.0, "B": 8.63, "D": 22.13, "E": 19.75, "F": 14.25, "H": 8.63}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    dv = DataValidation(type="list", formula1=f'"{STATUS_LIST}"', allow_blank=False)
    dv.add(f"F2:F{n_data_rows}")
    ws.add_data_validation(dv)


def build_changelog(wb):
    ws = wb.create_sheet("Changelog")
    for c, header in enumerate(CHANGELOG_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True)


def build_dashboard(wb, wards):
    ws = wb.create_sheet("Dashboard")

    for c, header in enumerate(DASHBOARD_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True)
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 37.0
    ws.column_dimensions["B"].width = 12.63
    ws.column_dimensions["G"].width = 15.38
    ws.column_dimensions["J"].width = 19.25
    ws.column_dimensions["L"].width = 19.25

    n = len(wards)
    first_row = 2
    last_ward_row = first_row + n - 1
    overall_row = last_ward_row + 1

    for i, (ward_name, _district) in enumerate(wards):
        r = first_row + i
        ws.cell(row=r, column=1, value=ward_name)
        ws.cell(row=r, column=2, value=f"=COUNTIF(Data!$D:$D,A{r})")
        ws.cell(row=r, column=3, value=f'=COUNTIFS(Data!$D:$D,A{r},Data!$F:$F,"Complete")')
        ws.cell(row=r, column=4, value=f'=COUNTIFS(Data!$D:$D,A{r},Data!$F:$F,"In_Progress")')
        ws.cell(row=r, column=5, value=f'=COUNTIFS(Data!$D:$D,A{r},Data!$F:$F,"Planned")')
        ws.cell(row=r, column=6, value=f'=COUNTIFS(Data!$D:$D,A{r},Data!$F:$F,"Not_Started")')
        ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)").number_format = "0.00%"
        ws.cell(row=r, column=8, value=f"=IFERROR(SUMIF(Data!$D:$D,A{r},Data!$G:$G),0)")
        ws.cell(row=r, column=9, value=(
            f'=IFERROR(SUMIFS(Data!$G:$G,Data!$D:$D,A{r},Data!$F:$F,"Complete")'
            f'+SUMIFS(Data!$G:$G,Data!$D:$D,A{r},Data!$F:$F,"In_Progress"),0)'
        ))
        ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/H{r},0)").number_format = "0.00%"
        # column 11 (Overseer) intentionally left blank — filled in manually per ward
        ws.cell(row=r, column=12, value=f"=H{r}-I{r}").number_format = "#,##0"

    r = overall_row
    ws.cell(row=r, column=1, value="Overall").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"=SUM(B{first_row}:B{last_ward_row})")
    ws.cell(row=r, column=3, value=f"=SUM(C{first_row}:C{last_ward_row})")
    ws.cell(row=r, column=4, value=f"=SUM(D{first_row}:D{last_ward_row})")
    ws.cell(row=r, column=5, value=f"=SUM(E{first_row}:E{last_ward_row})")
    ws.cell(row=r, column=6, value=f"=SUM(F{first_row}:F{last_ward_row})")
    ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)").number_format = "0.00%"
    ws.cell(row=r, column=8, value=f"=SUM(H{first_row}:H{last_ward_row})")
    ws.cell(row=r, column=9, value=f"=SUM(I{first_row}:I{last_ward_row})")
    ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/H{r},0)").number_format = "0.00%"
    ws.cell(row=r, column=12, value=f"=SUM(L{first_row}:L{last_ward_row})").number_format = "#,##0"

    for col in "ABCDEFGHIJL":
        ws.cell(row=r, column=ws[f"{col}1"].column).font = Font(bold=True)

    # ── Colour scales ──────────────────────────────────────────────────────
    ws.conditional_formatting.add(
        f"G{first_row}:G{last_ward_row}",
        ColorScaleRule(
            start_type="percent", start_value=0, start_color="FFFFFFFF",
            mid_type="percent", mid_value=50, mid_color="FFABDDC5",
            end_type="percent", end_value=100, end_color="FF57BB8A",
        ),
    )
    ws.conditional_formatting.add(
        f"J{first_row}:J{r}",
        ColorScaleRule(start_type="min", start_color="FFFFFFFF", end_type="max", end_color="FF57BB8A"),
    )
    ws.conditional_formatting.add(
        f"L{first_row}:L{last_ward_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFFFF", end_type="max", end_color="FFE06666"),
    )

    # ── Borough Summary KPI block ───────────────────────────────────────────
    kpi_title_row = overall_row + 2
    ws.cell(row=kpi_title_row, column=1, value="Borough Summary").font = Font(bold=True, size=12)

    kpi_rows = {}
    kpi_defs = [
        ("Total Roads",                                  f"=B{overall_row}",                                  "#,##0"),
        ("Total Residences (estimate)",                  f"=H{overall_row}",                                  "#,##0"),
        ("Residences Reached (Complete)",                 f'=SUMIFS(Data!$G:$G,Data!$F:$F,"Complete")',        "#,##0"),
        ("Residences Reached (incl. In Progress, 30% credit)",
         None,  # filled in below once its own row number is known (self-reference)
         "#,##0"),
        ("Residences Remaining",                          None, "#,##0"),
        ("% Residences Reached (Complete only)",          None, "0.00%"),
        ("Wards Fully Complete (100% of roads)",          f"=COUNTIF(G{first_row}:G{last_ward_row},1)",        "#,##0"),
        ("Wards Not Yet Started (0% of roads)",           f"=COUNTIF(G{first_row}:G{last_ward_row},0)",        "#,##0"),
    ]

    row = kpi_title_row + 1
    label_to_row = {}
    for label, formula, fmt in kpi_defs:
        ws.cell(row=row, column=1, value=label)
        if formula is not None:
            cell = ws.cell(row=row, column=2, value=formula)
            cell.number_format = fmt
        label_to_row[label] = row
        row += 1

    reached_complete_row = label_to_row["Residences Reached (Complete)"]
    reached_weighted_row = label_to_row["Residences Reached (incl. In Progress, 30% credit)"]
    remaining_row = label_to_row["Residences Remaining"]
    pct_reached_row = label_to_row["% Residences Reached (Complete only)"]
    total_residences_row = label_to_row["Total Residences (estimate)"]

    ws.cell(row=reached_weighted_row, column=2, value=(
        f"=B{reached_complete_row}"
        f'+0.3*SUMIFS(Data!$G:$G,Data!$F:$F,"In_Progress")'
    )).number_format = "#,##0"
    ws.cell(row=remaining_row, column=2, value=f"=B{total_residences_row}-B{reached_complete_row}").number_format = "#,##0"
    ws.cell(row=pct_reached_row, column=2, value=f"=IFERROR(B{reached_complete_row}/B{total_residences_row},0)").number_format = "0.00%"

    kpi_end_row = row - 1

    # ── Charts ───────────────────────────────────────────────────────────
    chart_row = kpi_end_row + 3

    area = AreaChart()
    area.title = "Leafletting % Complete by Ward"
    area.width, area.height = 22, 9
    cats = Reference(ws, min_col=1, min_row=first_row, max_row=overall_row)
    for col in (7, 10):
        data = Reference(ws, min_col=col, min_row=1, max_row=overall_row)
        area.add_data(data, titles_from_data=True)
    area.set_categories(cats)
    ws.add_chart(area, f"A{chart_row}")

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "percentStacked"
    bar.overlap = 100
    bar.title = "Leafletting Status by Ward"
    bar.width, bar.height = 22, 9
    for col in (5, 4, 6, 3):  # Planned, In Progress, Not Started, Complete
        data = Reference(ws, min_col=col, min_row=1, max_row=overall_row)
        bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, f"L{chart_row}")

    chart_row_2 = chart_row + 18

    pie = PieChart()
    pie.title = "Borough-Wide Status Breakdown"
    pie.width, pie.height = 12, 9
    pie_data = Reference(ws, min_col=3, max_col=6, min_row=overall_row)
    pie_cats = Reference(ws, min_col=3, max_col=6, min_row=1, max_row=1)
    pie.add_data(pie_data, titles_from_data=False, from_rows=True)
    pie.set_categories(pie_cats)
    ws.add_chart(pie, f"A{chart_row_2}")

    remaining_bar = BarChart()
    remaining_bar.type = "bar"
    remaining_bar.title = "Residences Remaining by Ward"
    remaining_bar.width, remaining_bar.height = 22, 9
    remaining_data = Reference(ws, min_col=12, min_row=1, max_row=last_ward_row)
    remaining_cats = Reference(ws, min_col=1, min_row=first_row, max_row=last_ward_row)
    remaining_bar.add_data(remaining_data, titles_from_data=True)
    remaining_bar.set_categories(remaining_cats)
    remaining_bar.legend = None
    ws.add_chart(remaining_bar, f"L{chart_row_2}")

    return reached_complete_row


def build_checksum(wb, reached_complete_row):
    ws = wb.create_sheet("Checksum")
    ws["A1"] = (
        '=COUNTIF(Data!F:F,"Not_Started")&"|"&COUNTIF(Data!F:F,"Planned")&"|"'
        '&COUNTIF(Data!F:F,"In_Progress")&"|"&COUNTIF(Data!F:F,"Complete")&"|"'
        f"&Dashboard!B{reached_complete_row}"
    )


def build_authorised(wb):
    ws = wb.create_sheet("Authorised")
    emails = cfg.get("authorised_emails") or []
    for i, email in enumerate(emails, start=1):
        ws.cell(row=i, column=1, value=email)


def main():
    if not Path(EXCEL_INPUT).exists():
        raise SystemExit(f"ERROR: {EXCEL_INPUT} not found — run estimate_residences_uprn.py first.")

    wb = load_workbook(EXCEL_INPUT)
    data_ws = wb["Data"]
    n_data_rows = data_ws.max_row

    format_data_sheet(data_ws, n_data_rows)
    build_changelog(wb)
    reached_complete_row = build_dashboard(wb, cfg["wards"])
    build_checksum(wb, reached_complete_row)
    build_authorised(wb)

    wb.save(EXCEL_OUTPUT)

    print(f"Wrote {EXCEL_OUTPUT}")
    print(f"  Data rows: {n_data_rows - 1:,}")
    print(f"  Wards on Dashboard: {len(cfg['wards'])}")
    print(f"  Authorised emails: {cfg.get('authorised_emails') or '(none set)'}")
    print("\nSheets: Data, Changelog, Dashboard, Checksum, Authorised")
    print("Ready to upload to Google Sheets.")


if __name__ == "__main__":
    main()
