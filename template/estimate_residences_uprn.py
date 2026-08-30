"""
estimate_residences_uprn.py  (TEMPLATE)
=========================================
Populates the 'Residences' column using OS Open UPRN address points matched
directly to road centrelines. Run this AFTER run_pipeline.py. Filenames are
derived automatically from constituency_config.py's output_prefix — no
manual edits needed.

WHY UPRN-BUFFER MATCHING, NOT OSM BUILDING FOOTPRINTS:
  An earlier approach assigned OSM building footprints (filtered to
  residential building=* tags) to the nearest road, using UPRNs only to
  size apartment blocks. It undercounts badly wherever OSM buildings are
  tagged generically as building=yes instead of house/detached/etc (very
  common) — see _archive/flawed_osm_building_method/README.md for the
  measured ~8x undercount on a real run. UPRNs are actual Royal Mail/OS
  address points and don't depend on OSM tagging quality at all.

Requirements: geopandas, pandas, shapely, openpyxl, pyproj

Inputs:
    <prefix>_Leafletting.xlsx        (from run_pipeline.py)
    <prefix>_constituency.geojson    (from run_pipeline.py)
    ../data/osopenuprn_202605.csv

Output:
    <prefix>_Leafletting_residences_uprn.xlsx
"""

import re
import time
import shutil
import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import LineString, MultiLineString
from openpyxl import load_workbook

from constituency_config import get_config

_cfg = get_config()
_prefix = _cfg["output_prefix"]

UPRN_CSV     = "../data/osopenuprn_202605.csv"
EXCEL_INPUT  = f"{_prefix}_Leafletting.xlsx"
EXCEL_OUTPUT = f"{_prefix}_Leafletting_residences_uprn.xlsx"
CONSTITUENCY = f"{_prefix}_constituency.geojson"

ROAD_BUFFER_METRES = 40
COMMERCIAL_CLUSTER_THRESHOLD = 150
UPRN_CHUNK_SIZE = 500_000


def parse_linestrings(geom_str):
    if pd.isna(geom_str) or not str(geom_str).strip():
        return None
    segments = []
    for part in str(geom_str).split("|"):
        m = re.match(r'LINESTRING\((.+)\)', part.strip())
        if not m:
            continue
        coords = []
        for pair in m.group(1).split(","):
            pts = pair.strip().split()
            if len(pts) == 2:
                try:
                    coords.append((float(pts[0]), float(pts[1])))
                except ValueError:
                    continue
        if len(coords) >= 2:
            segments.append(LineString(coords))
    if not segments:
        return None
    return segments[0] if len(segments) == 1 else MultiLineString(segments)


def main():
    t0 = time.time()

    print("Step 1: Loading constituency boundary...")
    boundary = gpd.read_file(CONSTITUENCY)
    constituency_boundary = boundary.geometry.iloc[0]
    constituency_padded = constituency_boundary.buffer(0.002)
    bbox = constituency_padded.bounds
    print(f"  Boundary bounds: {tuple(round(x, 4) for x in bbox)}")

    print(f"\nStep 2: Filtering UPRNs to constituency boundary...")
    print(f"  Reading '{UPRN_CSV}' in chunks of {UPRN_CHUNK_SIZE:,} rows...")

    kept, total_read, chunk_num = [], 0, 0

    for chunk in pd.read_csv(
        UPRN_CSV,
        chunksize=UPRN_CHUNK_SIZE,
        usecols=['UPRN', 'LATITUDE', 'LONGITUDE'],
        dtype={'UPRN': 'int64', 'LATITUDE': 'float64', 'LONGITUDE': 'float64'}
    ):
        chunk_num += 1
        total_read += len(chunk)

        pre = chunk[
            chunk['LONGITUDE'].between(bbox[0], bbox[2]) &
            chunk['LATITUDE'].between(bbox[1], bbox[3])
        ]

        if len(pre) > 0:
            pts = gpd.GeoDataFrame(pre, geometry=gpd.points_from_xy(pre['LONGITUDE'], pre['LATITUDE']), crs='EPSG:4326')
            inside = pts[pts.geometry.within(constituency_padded)]
            kept.append(inside[['UPRN', 'LATITUDE', 'LONGITUDE', 'geometry']])

        if chunk_num % 10 == 0:
            print(f"  ... {total_read:,} rows read, {sum(len(k) for k in kept):,} kept [{time.time()-t0:.0f}s]")

    uprn_gdf = pd.concat(kept, ignore_index=True)
    print(f"\n  Done. {len(uprn_gdf):,} UPRNs in constituency (from {total_read:,} GB-wide)")

    print(f"\n  Filtering commercial clusters (>={COMMERCIAL_CLUSTER_THRESHOLD} UPRNs at same coordinate)...")
    coord_key = uprn_gdf['LATITUDE'].round(5).astype(str) + ',' + uprn_gdf['LONGITUDE'].round(5).astype(str)
    coord_counts = coord_key.map(coord_key.value_counts())
    commercial_mask = coord_counts >= COMMERCIAL_CLUSTER_THRESHOLD
    n_filtered = commercial_mask.sum()
    uprn_gdf = uprn_gdf[~commercial_mask].copy()
    print(f"  Removed {n_filtered:,} likely commercial UPRNs, {len(uprn_gdf):,} residential remain")

    print(f"\nStep 3: Loading roads and buffering ({ROAD_BUFFER_METRES}m)...")
    df = pd.read_excel(EXCEL_INPUT, sheet_name="Data")

    raw_geoms = []
    for _, row in df.iterrows():
        raw_geoms.append(parse_linestrings(row.get('road_geometry')))

    roads_gdf = gpd.GeoDataFrame(df.copy(), geometry=raw_geoms, crs='EPSG:4326')
    roads_valid = roads_gdf[roads_gdf.geometry.notna()].copy()

    roads_proj = roads_valid.to_crs('EPSG:27700')
    uprn_proj = gpd.GeoDataFrame(uprn_gdf, geometry='geometry', crs='EPSG:4326').to_crs('EPSG:27700')

    roads_buffered = roads_proj.copy()
    roads_buffered['geometry'] = roads_proj.geometry.buffer(ROAD_BUFFER_METRES)
    print(f"  {len(roads_valid)} roads with geometry")

    print(f"\nStep 4: Spatial join - UPRNs to road buffers...")
    joined = gpd.sjoin(
        uprn_proj[['UPRN', 'geometry']],
        roads_buffered[['geometry']].reset_index().rename(columns={'index': 'road_idx'}),
        how='left',
        predicate='within'
    )

    multi_matched = joined[joined.index.duplicated(keep=False) & joined['road_idx'].notna()]
    n_multi = joined.index.duplicated(keep=False).sum()

    if n_multi > 0:
        print(f"  {n_multi:,} UPRNs matched multiple roads - resolving to nearest...")
        resolved = {}
        for uprn_i, group in multi_matched.groupby(level=0):
            pt = uprn_proj.loc[uprn_i, 'geometry']
            best_road = min(group['road_idx'], key=lambda ri: pt.distance(roads_proj.loc[ri, 'geometry']))
            resolved[uprn_i] = best_road

        # Keep exactly one row per UPRN (preserving its UPRN/geometry columns),
        # then overwrite that row's road_idx with the resolved nearest road.
        joined = joined[~joined.index.duplicated(keep='first')].copy()
        for uprn_i, road_idx in resolved.items():
            joined.loc[uprn_i, 'road_idx'] = road_idx

    orphans = joined[joined['road_idx'].isna()]
    print(f"  {len(orphans):,} orphaned UPRNs - assigning to nearest road...")

    if len(orphans) > 0:
        orphan_pts = uprn_proj.loc[orphans.index, ['UPRN', 'geometry']]
        sindex = roads_proj.sindex

        for uprn_i, row in orphan_pts.iterrows():
            pt = row['geometry']
            result = sindex.nearest(pt, return_all=False)
            if result.shape[1] > 0:
                tree_pos = result[1][0]
                nearest_ri = roads_proj.index[tree_pos]
                joined.loc[uprn_i, 'road_idx'] = nearest_ri

    matched = joined[joined['road_idx'].notna()].copy()
    matched['road_idx'] = matched['road_idx'].astype(int)
    counts = matched.groupby('road_idx')['UPRN'].count()

    print(f"\n  UPRNs assigned: {counts.sum():,}")
    print(f"  Roads with >=1 UPRN: {(counts > 0).sum()} of {len(roads_valid)}")

    print(f"\nStep 5: Writing results to '{EXCEL_OUTPUT}'...")
    shutil.copy(EXCEL_INPUT, EXCEL_OUTPUT)
    wb = load_workbook(EXCEL_OUTPUT)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    resid_col = headers.index('Residences') + 1

    for road_idx, count in counts.items():
        ws.cell(row=int(road_idx) + 2, column=resid_col).value = int(count)

    for idx in roads_valid.index:
        if idx not in counts.index:
            cell = ws.cell(row=int(idx) + 2, column=resid_col)
            if cell.value is None or str(cell.value) == '-':
                cell.value = 0

    wb.save(EXCEL_OUTPUT)

    elapsed = time.time() - t0
    print(f"\nComplete in {elapsed/60:.1f} minutes.")
    print(f"   Output: {EXCEL_OUTPUT}")
    print(f"  - {n_filtered:,} UPRNs at dense coordinate clusters were excluded as")
    print("    likely commercial/industrial.")


if __name__ == "__main__":
    main()
